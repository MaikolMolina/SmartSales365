from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

class ExcelGenerator:
    """
    Genera reportes en formato Excel
    """
    
    def generate_excel(self, data, parsed_command, query_results):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Reporte {parsed_command['report_type']}"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Reporte de {parsed_command['report_type'].title()}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = center_align
        
        # Información del reporte
        ws['A3'] = "Comando:"
        ws['B3'] = parsed_command['original_command']
        ws['A4'] = "Generado:"
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Tabla de datos
        if query_results:
            self._add_data_table(ws, query_results, header_font, header_fill, center_align)
        else:
            ws['A6'] = "No se encontraron datos para el reporte"
        
        # Ajustar anchos de columna
        self._adjust_column_widths(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _add_data_table(self, worksheet, query_results, header_font, header_fill, center_align):
        if not query_results:
            return
        
        # Encabezados
        headers = list(query_results[0].keys())
        start_row = 6
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Datos
        for row_idx, row_data in enumerate(query_results, start_row + 1):
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = row_data[header]
                cell.alignment = center_align
    
    def _adjust_column_widths(self, worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_excel_response(self, data, parsed_command, query_results, filename=None):
        excel_buffer = self.generate_excel(data, parsed_command, query_results)
        
        if not filename:
            filename = f"reporte_{parsed_command['report_type']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response